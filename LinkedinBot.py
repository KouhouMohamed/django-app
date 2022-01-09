import fonctions
import requests
from bs4 import BeautifulSoup
import openpyxl
from secrets import username, password
import sqlite3
        

if __name__ == '__main__':
    db = sqlite3.connect("ProjectStage\db.sqlite3")
    # Enmplacement de fichier xlsx ou le résultat est stocker 
    path_xlsx = r".\ListOfStudents.xlsx"
    work_book = openpyxl.load_workbook(path_xlsx)
    sheet = work_book.active
    row = 1
    colum = 1
    # first row in the table
    sheet.cell(row=row, column = colum).value = "Student"
    sheet.cell(row=row, column = colum + 1).value = "School"
    sheet.cell(row=row, column = colum + 2).value = "Image"
    
    #Extract students names from PDF file
    ListOfStudents = fonctions.find_name()

    #create an objects from the class LinkedinBot 
    bot = fonctions.LinkedinBot(username, password)

    #Login to Linkedin with the inputs informations
    bot.login(username, password)

    # looking for the school foe each student
    for student in ListOfStudents:
        # go to the next row to fill it
        row = row + 1
        liste = bot.search(student)
        if len(liste) == 0:
            bot.school = 'account not found'
            bot.image = 'account not found'
            bot._nav(bot.feed_url)
        else:
            url = liste[0]
            bot.searchSchool(url)
        # fill the row with the name of student and his school names
        sheet.cell(row=row, column = colum).value = student
        sheet.cell(row=row, column = colum + 1).value = bot.school
        sheet.cell(row=row, column = colum + 2).value = bot.image
        db.execute("insert into studentschool_studentstage(name,school,image) values (?,?,?)", (student, bot.school,bot.image))
        db.commit()
        #save the work
        work_book.save(path_xlsx)
    
    students = db.execute("select name from studentschool_studentstage where school in ('None','account not found')")
    for student in students:
        url_req = requests.get('https://google.com/search?q=' + student[0])
        soup = BeautifulSoup(url_req.text, 'html.parser')
        div_container = soup.findAll("div", {"class": "kCrYT"})
        for div in div_container:
            if div.a and ( div.a.text.find("linkedin.com") > 0):
                bot.searchSchool(fonctions.find_link(div.a["href"]))
                db.execute("update studentschool_studentstage set school={},image={} where name={}".format(bot.school,bot.image,student[0]))
    

    bot.quit()